#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
手车互联PM提效工具 - Web 版
运行后自动打开浏览器，输入 Jira ID 即可查询状态
账号密码自动读取 config.json，无需在页面输入
支持钉钉机器人一键批量提醒未关闭 Bug
"""

import json
import urllib.request
import urllib.error
import urllib.parse
import base64
import re
import os
import sys
import threading
import webbrowser
import socketserver
from http.server import BaseHTTPRequestHandler, HTTPServer

# ─────────────────────────────────────────────
# 配置
# ─────────────────────────────────────────────
PORT = 8890

# 脚本所在目录（PyInstaller 打包后 sys._MEIPASS 为临时解压目录）
if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_FILE = os.path.join(APP_DIR, "config.json")

DEFAULT_CONFIG = {
    "jira_base_url": "https://jira.gwm.cn",
    "jira_username": "",
    "jira_password": "",
    "dingtalk_webhook": "",   # 钉钉自定义机器人 Webhook URL
}

STATUS_MAP = {
    "待分析": "待分析",
    "分析中": "分析中",
    "修正中": "修正中",
    "修正完成": "修正完成",
    "待验证": "待验证",
    "验证中": "验证中",
    "完成": "完成",
    "关闭": "关闭",
    "重新打开": "重新打开",
    "Open": "Open（待处理）",
    "In Progress": "进行中",
    "Resolved": "已解决",
    "Closed": "已关闭",
    "Reopened": "重新打开",
    "Done": "完成",
}

# 状态对应的颜色标签（前端用）
STATUS_COLOR = {
    "待分析": "gray",
    "分析中": "blue",
    "修正中": "orange",
    "修正完成": "green",
    "待验证": "purple",
    "验证中": "purple",
    "完成": "green",
    "关闭": "green",
    "重新打开": "red",
    "Open（待处理）": "gray",
    "进行中": "blue",
    "已解决": "green",
    "已关闭": "green",
    "Done": "green",
}

# 已完成状态（除此之外都算未关闭，用于钉钉提醒过滤）
DONE_STATUSES = {"完成", "关闭", "已关闭", "已解决", "Done", "Closed", "Resolved"}


def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return dict(DEFAULT_CONFIG)


def make_auth_header(username: str, password: str) -> str:
    credentials = f"{username}:{password}"
    encoded = base64.b64encode(credentials.encode("utf-8")).decode("utf-8")
    return f"Basic {encoded}"


def map_status(raw_status: str) -> str:
    return STATUS_MAP.get(raw_status, raw_status)


def get_issue_status(issue_id: str, base_url: str, auth_header: str) -> dict:
    url = f"{base_url}/rest/api/2/issue/{issue_id}?fields=summary,status,assignee,priority,issuetype,resolution"
    req = urllib.request.Request(url)
    req.add_header("Authorization", auth_header)
    req.add_header("Content-Type", "application/json")
    req.add_header("Accept", "application/json")

    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            fields = data.get("fields", {})
            raw_status = fields.get("status", {}).get("name", "未知")
            status_cn = map_status(raw_status)
            assignee_obj = fields.get("assignee") or {}
            resolution_obj = fields.get("resolution") or {}
            priority_obj = fields.get("priority") or {}
            issue_type_obj = fields.get("issuetype") or {}
            return {
                "id": issue_id,
                "summary": fields.get("summary", ""),
                "status": status_cn,
                "status_raw": raw_status,
                "status_color": STATUS_COLOR.get(status_cn, "gray"),
                "assignee": assignee_obj.get("displayName", "未分配"),
                "assignee_key": assignee_obj.get("name", "") or assignee_obj.get("key", ""),
                "resolution": resolution_obj.get("name", "—"),
                "priority": priority_obj.get("name", "—"),
                "issue_type": issue_type_obj.get("name", "—"),
                "error": None,
            }
    except urllib.error.HTTPError as e:
        msg = {404: "ID 不存在（404）", 401: "认证失败（401）", 403: "无权限（403）"}.get(e.code, f"HTTP {e.code}")
        return {"id": issue_id, "error": msg}
    except urllib.error.URLError as e:
        return {"id": issue_id, "error": f"网络错误：{e.reason}"}
    except Exception as e:
        return {"id": issue_id, "error": str(e)}


def extract_employee_id(assignee: str) -> str:
    """从 Jira assignee displayName 中提取工号，如 '赵永洪GW00295409' → 'GW00295409'"""
    m = re.search(r'([A-Z]{2,}\d+)', assignee)
    return m.group(1) if m else ""


def build_markdown_text(results: list, base_url: str) -> tuple:
    """构建钉钉 Markdown 消息文本，返回 (text, open_count, assignee_count, assignees, at_user_ids)"""
    open_bugs = [r for r in results if not r.get("error") and r.get("status") not in DONE_STATUSES]
    if not open_bugs:
        return ("", 0, 0, [], [])

    by_assignee = {}
    for bug in open_bugs:
        assignee = bug.get("assignee", "未分配")
        if assignee not in by_assignee:
            by_assignee[assignee] = []
        by_assignee[assignee].append(bug)

    # 收集所有经办人的钉钉 userId（优先用 Jira username=工号，兜底正则）
    at_user_ids = []
    for assignee, bugs in by_assignee.items():
        # 优先从第一个 bug 的 assignee_key 取（即 Jira username）
        eid = bugs[0].get("assignee_key", "") or extract_employee_id(assignee)
        if eid:
            at_user_ids.append(eid)

    priority_icon = {
        "最高": "🔴", "High": "🔴", "P0": "🔴",
        "高": "🟠",   "P1": "🟠",
        "中": "🟡",   "Medium": "🟡", "P2": "🟡",
        "低": "🟢",   "Low": "🟢",   "P3": "🟢",
        "最低": "⚪",  "Lowest": "⚪",
    }

    lines = []
    lines.append(f"> 网友朋友们好，我是魏建军，以下为当前未关闭的缺陷，请相关经办人尽快处理，辛苦大家！\n")
    for assignee, bugs in sorted(by_assignee.items()):
        eid = bugs[0].get("assignee_key", "") or extract_employee_id(assignee)
        at_tag = f" @{eid}" if eid else ""
        lines.append(f"👤 {assignee}{at_tag}（{len(bugs)} 个）")
        for bug in bugs:
            p_icon = priority_icon.get(bug.get("priority", ""), "▪")
            bug_url = f"{base_url}/browse/{bug['id']}"
            lines.append(f"- {p_icon} [{bug['id']}]({bug_url})  `{bug['status']}`")
        lines.append("")

    lines.append("*该消息由Python本地服务点击发送，如有疑问请联系 赵永洪GW00295409*")

    return ("\n".join(lines), len(open_bugs), len(by_assignee), list(by_assignee.keys()), at_user_ids)


def send_dingtalk_notify(webhook_url: str, results: list, base_url: str, custom_text: str = None) -> dict:
    """
    发送钉钉群消息，精准 @对应经办人。
    custom_text: 如果提供则直接使用，否则根据 results 自动构建。
    """
    if custom_text:
        markdown_text = custom_text
        # 从文本中大致估算数量
        open_bugs = [r for r in results if not r.get("error") and r.get("status") not in DONE_STATUSES]
        open_count = len(open_bugs)
        by_assignee = {}
        at_user_ids = []
        for bug in open_bugs:
            assignee = bug.get("assignee", "未分配")
            if assignee not in by_assignee:
                by_assignee[assignee] = []
            by_assignee[assignee].append(bug)
            eid = bug.get("assignee_key", "") or extract_employee_id(assignee)
            if eid and eid not in at_user_ids:
                at_user_ids.append(eid)
        assignee_count = len(by_assignee)
        assignees = list(by_assignee.keys())
    else:
        markdown_text, open_count, assignee_count, assignees, at_user_ids = build_markdown_text(results, base_url)
        if not markdown_text:
            return {"success": False, "error": "没有未关闭的 Bug，无需提醒"}

    # 钉钉 markdown 消息中 @人 需要写 @工号
    markdown_text = "@" + " @".join(at_user_ids) + "\n" + markdown_text if at_user_ids else "以下 Bug 需处理\n" + markdown_text

    payload = {
        "msgtype": "markdown",
        "markdown": {
            "title": f"Jira 未关闭 Bug 提醒（{open_count} 个）",
            "text": markdown_text,
        },
        "at": {
            "atUserIds": at_user_ids,
            "isAtAll": False,
        },
    }

    try:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            webhook_url,
            data=body,
            method="POST"
        )
        req.add_header("Content-Type", "application/json; charset=utf-8")
        req.add_header("Content-Length", str(len(body)))

        with urllib.request.urlopen(req, timeout=15) as resp:
            resp_data = json.loads(resp.read().decode("utf-8"))
            if resp_data.get("errcode") == 0:
                return {
                    "success": True,
                    "open_count": open_count,
                    "assignee_count": assignee_count,
                    "assignees": assignees,
                }
            else:
                return {
                    "success": False,
                    "error": f"钉钉返回错误：{resp_data.get('errmsg', '未知错误')}（code={resp_data.get('errcode')}）"
                }
    except urllib.error.HTTPError as e:
        body_msg = e.read().decode("utf-8", errors="ignore")
        return {"success": False, "error": f"HTTP {e.code}：{body_msg[:200]}"}
    except urllib.error.URLError as e:
        return {"success": False, "error": f"网络错误：{e.reason}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


# ─────────────────────────────────────────────
# Huawei 链接检查 —— 后端函数
# ─────────────────────────────────────────────

def make_auth(username, password):
    cred = f"{username}:{password}"
    encoded = base64.b64encode(cred.encode("utf-8")).decode("utf-8")
    return f"Basic {encoded}"


def jira_api(path, base_url, auth_header):
    """发起 Jira API GET 请求"""
    url = f"{base_url.rstrip('/')}{path}"
    req = urllib.request.Request(url)
    req.add_header("Authorization", auth_header)
    req.add_header("Content-Type", "application/json")
    req.add_header("Accept", "application/json")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw)
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        return {"error": f"HTTP {e.code}: {body[:200]}"}
    except urllib.error.URLError as e:
        return {"error": f"网络错误: {e.reason}"}


def extract_huawei_links(text):
    """提取包含 huawei 的链接"""
    found = []
    for m in re.finditer(r'https?://[^\s\]"\'<>]+', text, re.IGNORECASE):
        url = m.group(0)
        if "huawei" in url.lower():
            found.append(url)
    for m in re.finditer(r'\[([^\]]*)\|([^\]]+)\]', text, re.IGNORECASE):
        label, url = m.group(1), m.group(2)
        if "huawei" in url.lower() or "huawei" in label.lower():
            found.append(url.strip())
    seen = set()
    return [u for u in found if not (u in seen or seen.add(u))]


def check_comments(issue_id, base_url, auth_header):
    """检查指定 Issue 的备注"""
    data = jira_api(f"/rest/api/2/issue/{issue_id}/comment?maxResults=500", base_url, auth_header)
    if "error" in data:
        return [], data["error"]
    comments = data.get("comments", [])
    hits = []
    for c in comments:
        body = c.get("body", "") or ""
        links = extract_huawei_links(body)
        if links:
            hits.append({
                "issue_id": issue_id,
                "comment_id": c.get("id", ""),
                "author": c.get("author", {}).get("displayName", "未知"),
                "created": c.get("created", "")[:19].replace("T", " "),
                "huawei_links": " | ".join(links),
                "comment_snippet": body[:200].replace("\r\n", " ").replace("\n", " "),
                "jira_url": f"{base_url}/browse/{issue_id}",
                "source": "直接命中",
                "parent_issue_id": "",
            })
    return hits, None


def get_linked_keys(issue_id, base_url, auth_header):
    """获取关联 Issue Key 列表"""
    data = jira_api(f"/rest/api/2/issue/{issue_id}?fields=issuelinks", base_url, auth_header)
    if "error" in data:
        return []
    keys = []
    for link in data.get("fields", {}).get("issuelinks", []):
        for direction in ("inwardIssue", "outwardIssue"):
            issue = link.get(direction)
            if issue:
                key = issue.get("key", "")
                if key and key not in keys:
                    keys.append(key)
    return keys


def check_issue(issue_id, base_url, auth_header, check_linked=True):
    """检查单个 Issue（含可选的关联检查）"""
    all_hits = []
    visited = set()

    hits, err = check_comments(issue_id, base_url, auth_header)
    if err:
        return {"error": err}
    all_hits.extend(hits)

    if not hits and check_linked:
        visited.add(issue_id)
        linked = get_linked_keys(issue_id, base_url, auth_header)
        for lk in linked:
            if lk in visited:
                continue
            visited.add(lk)
            lhits, _ = check_comments(lk, base_url, auth_header)
            for h in lhits:
                h["source"] = f"关联Issue (来自 {issue_id})"
                h["parent_issue_id"] = issue_id
            all_hits.extend(lhits)

    return {"hits": all_hits}


def build_huawei_csv(rows):
    """用纯 Python 构建 CSV 数据"""
    import csv
    from io import BytesIO
    buf = BytesIO()
    fieldnames = ["Issue ID", "Jira链接", "来源类型", "父Issue", "备注ID", "备注作者", "备注时间", "Huawei链接", "备注摘要(前200字)"]
    w = csv.writer(buf)
    w.writerow(fieldnames)
    for r in rows:
        w.writerow([r.get("issue_id"), r.get("jira_url"), r.get("source"), r.get("parent_issue_id"),
                     r.get("comment_id"), r.get("author"), r.get("created"), r.get("huawei_links"), r.get("comment_snippet")])
    return buf.getvalue(), "csv"


def build_huawei_xlsx(rows):
    """用 openpyxl 构建 xlsx"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "Huawei链接检查结果"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    headers = ["Issue ID", "Jira链接", "来源类型", "父Issue", "备注ID", "备注作者", "备注时间", "Huawei链接", "备注摘要(前200字)"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for r in rows:
        ws.append([r.get("issue_id"), r.get("jira_url"), r.get("source"), r.get("parent_issue_id"),
                    r.get("comment_id"), r.get("author"), r.get("created"), r.get("huawei_links"), r.get("comment_snippet")])

    for row in range(2, len(rows) + 2):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(name="微软雅黑", size=10)
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 50
    ws.column_dimensions["I"].width = 50
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(rows) + 1}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), "xlsx"


# ─────────────────────────────────────────────
# 评论转单分析 —— 后端函数
# ─────────────────────────────────────────────

ACTION_KEYWORDS = [
    "转手机侧分析", "转手机侧", "手机侧分析",
    "提单给手机侧", "提交给手机侧",
    "转给手机侧", "给手机侧分析",
    "转联盟分析", "提单给联盟", "提交给联盟",
    "转给联盟分析", "给联盟分析",
    "转荣耀手机侧分析", "提单给荣耀", "转给荣耀",
    "荣耀手机侧分析", "给荣耀分析",
    "转华为分析", "提单给华为", "转给华为",
    "华为分析", "提单华为", "转华为",
    "华为手机侧分析",
    "转手机端分析", "手机端分析",
    "转APP侧分析", "APP侧分析", "app侧分析",
    "转应用侧分析", "应用侧分析",
    "提单给手机", "提交手机侧",
    "已转手机侧", "已转联盟", "已转荣耀",
    "已提单给手机", "已提单给联盟", "已提单给荣耀",
    "已转华为", "已提单给华为",
    "流转至手机", "流转至联盟", "流转至荣耀",
    "流转到手机", "流转到联盟", "流转到荣耀",
    "分配给手机", "分配给联盟", "分配给荣耀",
    "移交手机侧", "移交联盟", "移交荣耀",
    "转手机团队", "转联盟团队", "转荣耀团队",
    "提交手机团队", "提交联盟团队", "提交荣耀团队",
    "协助分析", "协同分析", "配合分析",
    "需要手机侧确认", "需要联盟确认", "需要荣耀确认",
    "需要手机侧协助", "需要联盟协助", "需要荣耀协助",
    "请手机侧", "请联盟", "请荣耀",
    "建议转", "建议提单","hicar协助分析","联盟协助分析","手机协助分析","手机侧协助分析","荣耀协助分析","荣耀手机协助分析",
    "提单hicar","提单联盟","提单手机侧","提单荣耀手机侧","提单荣耀",
    "提单华为","hicar排查","联盟排查","荣耀排查","hicar分析","联盟分析","手机分析","手机侧分析","荣耀分析","荣耀手机分析",
    "手机端看下","手机端接力","同CPM单","hicar提单","华为提单",
    "华为继续定位","华为","联盟","荣耀","手机侧",
    "Hicar处理","hicar处理","手机处理","联盟处理","荣耀处理",
    "HiCar的同事","HiCar同事","hiCar的同事","hiCar同事","联盟的同事","联盟同事","荣耀同事","联盟的同事",
    "CEPM","提单","接力排查",
    "hicar看下","联盟看下","荣耀看下","手机侧看下"
]

def build_transfer_keyword_pattern():
    sorted_kw = sorted(ACTION_KEYWORDS, key=len, reverse=True)
    escaped = [re.escape(kw) for kw in sorted_kw]
    return re.compile("|".join(escaped), re.IGNORECASE)

TRANSFER_KEYWORD_PATTERN = build_transfer_keyword_pattern()


def strip_html(text: str) -> str:
    """移除 HTML 标签，返回纯文本"""
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    text = text.replace('&nbsp;', ' ').replace('&lt;', '<').replace('&gt;', '>')
    text = text.replace('&amp;', '&').replace('&quot;', '"')
    text = re.sub(r'\[([^\]|]*)\|([^\]]+)\]', r'\1(\2)', text)
    return text.strip()


def analyze_comment(body: str) -> list:
    """分析备注内容，返回命中的关键词列表"""
    plain_text = strip_html(body)
    matches = TRANSFER_KEYWORD_PATTERN.findall(plain_text)
    return list(dict.fromkeys(matches))


def get_transfer_comments(issue_id: str, base_url: str, auth_header: str) -> tuple:
    """获取指定 Issue 的所有备注，分析转单关键词，返回 (hits, error)"""
    data = jira_api(f"/rest/api/2/issue/{issue_id}/comment?maxResults=500&orderBy=-created", base_url, auth_header)
    if "error" in data:
        return [], data["error"]
    comments = data.get("comments", [])
    hits = []
    for c in comments:
        body = c.get("body", "") or ""
        matched = analyze_comment(body)
        if matched:
            author_obj = c.get("author", {})
            hits.append({
                "issue_id": issue_id,
                "comment_id": c.get("id", ""),
                "author_name": author_obj.get("name", "未知"),
                "author_display": author_obj.get("displayName", "未知"),
                "author_email": author_obj.get("emailAddress", ""),
                "created": c.get("created", "")[:19].replace("T", " "),
                "updated": c.get("updated", "")[:19].replace("T", " "),
                "matched_keywords": "、".join(matched),
                "comment_body_plain": strip_html(body),
                "jira_url": f"{base_url}/browse/{issue_id}",
            })
    return hits, None


def build_transfer_csv(rows: list):
    """纯 Python 构建 CSV"""
    import csv
    from io import BytesIO
    buf = BytesIO()
    fieldnames = ["Issue ID", "Jira链接", "备注ID", "备注时间", "更新时间",
                  "作者账号", "作者姓名", "作者邮箱", "匹配关键词", "备注内容(纯文本)"]
    w = csv.writer(buf)
    w.writerow(fieldnames)
    for r in rows:
        w.writerow([r["issue_id"], r["jira_url"], r["comment_id"], r["created"], r["updated"],
                     r["author_name"], r["author_display"], r["author_email"],
                     r["matched_keywords"], r["comment_body_plain"]])
    return buf.getvalue(), "csv"


def build_transfer_xlsx(rows: list):
    """用 openpyxl 构建 xlsx"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "转派分析结果"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    headers = ["Issue ID", "Jira链接", "备注ID", "备注时间", "更新时间",
               "作者账号", "作者姓名", "作者邮箱", "匹配关键词", "备注内容(纯文本)"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for r in rows:
        ws.append([r["issue_id"], r["jira_url"], r["comment_id"], r["created"], r["updated"],
                    r["author_name"], r["author_display"], r["author_email"],
                    r["matched_keywords"], r["comment_body_plain"]])

    for row in range(2, len(rows) + 2):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).font = Font(name="微软雅黑", size=10)
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=2).font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
        ws.cell(row=row, column=2).hyperlink = rows[row - 2]["jira_url"]

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 60
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{len(rows) + 1}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), "xlsx"


# ─────────────────────────────────────────────
# HTML 页面
# ─────────────────────────────────────────────
HTML_PAGE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>手车互联PM提效工具</title>
  <style>
    /* ═══════════════════════════════════════
       1. Reset & 基础样式
       ═══════════════════════════════════════ */
    * { box-sizing: border-box; margin: 0; padding: 0; }

    body {
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: #f0f2f5;
      color: #333;
      min-height: 100vh;
      padding: 24px 16px;
    }

    .container { max-width: 960px; margin: 0 auto; }

    /* ═══════════════════════════════════════
       2. 顶部 Header
       ═══════════════════════════════════════ */
    .header {
      background: linear-gradient(135deg, #1e3a5f 0%, #2d6cdf 100%);
      color: #fff;
      padding: 24px 28px;
      border-radius: 12px;
      margin-bottom: 20px;
      display: flex;
      align-items: center;
      gap: 14px;
    }
    .header-icon { font-size: 32px; }
    .header h1   { font-size: 20px; font-weight: 600; }
    .header p    { font-size: 13px; opacity: 0.8; margin-top: 4px; }

    /* ═══════════════════════════════════════
       3. 卡片容器
       ═══════════════════════════════════════ */
    .card {
      background: #fff;
      border-radius: 10px;
      padding: 24px;
      margin-bottom: 16px;
      box-shadow: 0 1px 4px rgba(0, 0, 0, 0.08);
    }

    /* ═══════════════════════════════════════
       4. 表单行（用户名 / 密码）
       ═══════════════════════════════════════ */
    .form-row {
      display: flex;
      gap: 12px;
      margin-bottom: 16px;
    }
    .form-group { flex: 1; }
    .form-group label {
      display: block;
      font-size: 13px;
      font-weight: 500;
      color: #555;
      margin-bottom: 5px;
    }
    .form-input {
      width: 100%;
      padding: 8px 12px;
      border: 1px solid #ddd;
      border-radius: 7px;
      font-size: 13px;
      outline: none;
      transition: border-color 0.2s;
    }
    .form-input:focus { border-color: #2d6cdf; }

    /* 密码框（含眼睛按钮） */
    .pwd-wrap  { position: relative; }
    .pwd-toggle {
      position: absolute;
      right: 8px;
      top: 50%;
      transform: translateY(-50%);
      background: none;
      border: none;
      cursor: pointer;
      font-size: 15px;
      color: #999;
      padding: 2px;
    }
    .pwd-toggle:hover { color: #555; }

    /* 记住密码 复选框 */
    .remember-row {
      display: flex;
      align-items: center;
      gap: 6px;
      margin-top: 8px;
      margin-bottom: 4px;
    }
    .remember-row input[type="checkbox"] {
      width: 15px;
      height: 15px;
      accent-color: #2d6cdf;
      cursor: pointer;
    }
    .remember-row label {
      font-size: 12px;
      color: #888;
      cursor: pointer;
      user-select: none;
      display: flex;
      align-items: center;
      gap: 4px;
    }
    .remember-row label:hover { color: #555; }
    .remember-row .clear-saved {
      font-size: 11px;
      color: #e74c3c;
      cursor: pointer;
      text-decoration: underline;
      margin-left: 8px;
      display: none;
    }
    .remember-row .clear-saved.visible { display: inline; }

    /* ═══════════════════════════════════════
       5. ID 输入区
       ═══════════════════════════════════════ */
    .input-label {
      font-size: 14px;
      font-weight: 500;
      color: #555;
      margin-bottom: 8px;
      display: block;
    }
    .input-hint {
      font-size: 12px;
      color: #999;
      margin-bottom: 10px;
    }
    textarea {
      width: 100%;
      height: 120px;
      border: 1px solid #ddd;
      border-radius: 8px;
      padding: 10px 12px;
      font-size: 14px;
      font-family: "SFMono-Regular", Consolas, monospace;
      resize: vertical;
      outline: none;
      transition: border-color 0.2s;
      line-height: 1.6;
    }
    textarea:focus { border-color: #2d6cdf; }

    /* ═══════════════════════════════════════
       6. 按钮
       ═══════════════════════════════════════ */
    .btn-row { display: flex; gap: 10px; margin-top: 14px; flex-wrap: wrap; }
    .btn {
      padding: 9px 22px;
      border: none;
      border-radius: 7px;
      font-size: 14px;
      font-weight: 500;
      cursor: pointer;
      transition: background 0.2s, transform 0.1s;
    }
    .btn:active { transform: scale(0.97); }

    .btn-primary {
      background: #2d6cdf;
      color: #fff;
    }
    .btn-primary:hover  { background: #1e5abf; }
    .btn-primary:disabled {
      background: #a0b8e8;
      cursor: not-allowed;
      transform: none;
    }

    .btn-secondary {
      background: #f0f2f5;
      color: #555;
    }
    .btn-secondary:hover { background: #e0e3e8; }

    .btn-export {
      background: #18a058;
      color: #fff;
      display: none;
    }
    .btn-export:hover { background: #129048; }
    .btn-export:disabled { background: #7cc4a0; cursor: not-allowed; transform: none; }

    /* 钉钉通知按钮 */
    .btn-dingtalk {
      background: #1aad19;
      color: #fff;
      opacity: 0.55;
    }
    .btn-dingtalk:not(:disabled) { opacity: 1; cursor: pointer; }
    .btn-dingtalk:not(:disabled):hover { background: #129811; }
    .btn-dingtalk:disabled { background: #7ec97e; cursor: not-allowed; transform: none; }

    /* 消息编辑区 */
    .edit-area-wrap { display: none; margin-top: 12px; }
    .edit-area-wrap .edit-label {
      font-size: 12px;
      font-weight: 500;
      color: #666;
      margin-bottom: 6px;
      display: flex;
      align-items: center;
      gap: 6px;
    }
    .edit-msg-textarea {
      width: 100%;
      min-height: 200px;
      max-height: 400px;
      border: 1px solid #ddd;
      border-radius: 8px;
      padding: 10px 12px;
      font-size: 12px;
      font-family: "SFMono-Regular", Consolas, monospace;
      line-height: 1.6;
      resize: vertical;
      outline: none;
      background: #fafbfd;
      color: #333;
    }
    .edit-msg-textarea:focus { border-color: #2d6cdf; background: #fff; }
    .edit-hint {
      font-size: 11px;
      color: #999;
      margin-top: 4px;
    }

    /* ═══════════════════════════════════════
       7. 进度条
       ═══════════════════════════════════════ */
    .progress-wrap { display: none; margin-top: 14px; }
    .progress-bar-bg {
      background: #e8edf5;
      border-radius: 99px;
      height: 8px;
      overflow: hidden;
    }
    .progress-bar-fill {
      height: 100%;
      background: linear-gradient(90deg, #2d6cdf, #5b9cf6);
      border-radius: 99px;
      transition: width 0.3s ease;
      width: 0%;
    }
    .progress-text {
      font-size: 12px;
      color: #888;
      margin-top: 6px;
    }

    /* ═══════════════════════════════════════
       8. 统计标签
       ═══════════════════════════════════════ */
    .stats-row {
      display: flex;
      gap: 12px;
      flex-wrap: wrap;
      margin-bottom: 16px;
    }
    .stat-chip {
      display: flex;
      align-items: center;
      gap: 6px;
      background: #f5f7fa;
      border-radius: 20px;
      padding: 5px 14px;
      font-size: 13px;
      color: #444;
    }
    .stat-chip .dot {
      width: 8px;
      height: 8px;
      border-radius: 50%;
    }
    .dot-total  { background: #2d6cdf; }
    .dot-done   { background: #18a058; }
    .dot-open   { background: #e6a23c; }
    .dot-error  { background: #e74c3c; }
    .dot-notify { background: #1aad19; }

    /* ═══════════════════════════════════════
       9. 结果表格
       ═══════════════════════════════════════ */
    #result-section { display: none; }

    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }
    thead tr { background: #f5f7fa; }

    th {
      text-align: left;
      padding: 10px 12px;
      font-weight: 600;
      color: #555;
      border-bottom: 2px solid #e8ecf0;
      white-space: nowrap;
    }
    td {
      padding: 10px 12px;
      border-bottom: 1px solid #f0f2f5;
      vertical-align: middle;
    }
    tr:hover td { background: #fafbfd; }

    /* ID 超链接 */
    .id-link {
      color: #2d6cdf;
      text-decoration: none;
      font-weight: 600;
      font-family: monospace;
    }
    .id-link:hover { text-decoration: underline; }

    /* 摘要单元格（超长截断） */
    .summary-cell {
      max-width: 300px;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
      color: #555;
    }

    /* ═══════════════════════════════════════
       10. 状态徽章
       ═══════════════════════════════════════ */
    .badge {
      display: inline-block;
      padding: 3px 10px;
      border-radius: 20px;
      font-size: 12px;
      font-weight: 500;
      white-space: nowrap;
    }
    .badge-green  { background: #e8f8ef; color: #18a058; }
    .badge-blue   { background: #e8f0fd; color: #2d6cdf; }
    .badge-orange { background: #fef3e2; color: #e6a23c; }
    .badge-gray   { background: #f0f2f5; color: #888; }
    .badge-red    { background: #fdecea; color: #e74c3c; }
    .badge-purple { background: #f0eaff; color: #7c4ddd; }
    .badge-error  { background: #fdecea; color: #e74c3c; }

    /* ═══════════════════════════════════════
       11. 日志面板
       ═══════════════════════════════════════ */
    .log-wrap {
      margin-top: 12px;
      background: #1a1e2e;
      border-radius: 8px;
      padding: 12px 14px;
      max-height: 160px;
      overflow-y: auto;
      display: none;
    }
    .log-line {
      font-size: 12px;
      font-family: monospace;
      line-height: 1.7;
      color: #8bb8e8;
    }
    .log-ok   { color: #5dbf85; }
    .log-err  { color: #f07070; }
    .log-info { color: #aaa; }
    .log-warn { color: #f0c040; }

    /* ═══════════════════════════════════════
       12. 空状态
       ═══════════════════════════════════════ */
    .empty-state {
      text-align: center;
      color: #bbb;
      padding: 40px 0;
      font-size: 14px;
    }

    /* ═══════════════════════════════════════
       13. 钉钉配置区
       ═══════════════════════════════════════ */
    .dingtalk-section {
      border: 1px solid #e8edf5;
      border-radius: 8px;
      padding: 16px;
      margin-top: 16px;
      background: #fafbfd;
    }
    .dingtalk-section .section-title {
      font-size: 13px;
      font-weight: 600;
      color: #444;
      margin-bottom: 10px;
      display: flex;
      align-items: center;
      gap: 6px;
    }
    .webhook-row {
      display: flex;
      gap: 8px;
      align-items: center;
    }
    .webhook-row .form-input {
      flex: 1;
      font-size: 12px;
      font-family: monospace;
    }
    .btn-save-webhook {
      padding: 7px 14px;
      background: #4a90d9;
      color: #fff;
      border: none;
      border-radius: 7px;
      font-size: 12px;
      cursor: pointer;
      white-space: nowrap;
    }
    .btn-save-webhook:hover { background: #3a7cc0; }

    .webhook-hint {
      font-size: 11px;
      color: #aaa;
      margin-top: 6px;
      line-height: 1.6;
    }

    /* 钉钉通知结果提示 */
    .notify-result {
      display: none;
      margin-top: 10px;
      padding: 10px 14px;
      border-radius: 7px;
      font-size: 13px;
      line-height: 1.6;
    }
    .notify-result.success {
      background: #e8f8ef;
      color: #18a058;
      border: 1px solid #b2dfc6;
    }
    .notify-result.error {
      background: #fdecea;
      color: #c0392b;
      border: 1px solid #f5b7b1;
    }

    /* ===== 工具导航栏 ===== */
    .tool-nav {
      display: flex;
      gap: 14px;
      margin-bottom: 24px;
      flex-wrap: wrap;
      justify-content: center;
      padding: 14px 20px;
      background: #fff;
      border-radius: 14px;
      box-shadow: 0 2px 12px rgba(0,0,0,.08);
    }
    .tool-nav-btn {
      padding: 11px 28px;
      border: 2px solid transparent;
      border-radius: 12px;
      font-size: 15px;
      font-weight: 600;
      text-decoration: none;
      cursor: pointer;
      transition: all .25s;
      display: inline-flex;
      align-items: center;
      gap: 8px;
      background: #f2f4f8;
      color: #4a5568;
      box-shadow: 0 1px 3px rgba(0,0,0,.05);
    }
    .tool-nav-btn:hover {
      border-color: #4472c4;
      color: #4472c4;
      background: #eef3ff;
      box-shadow: 0 3px 12px rgba(68,114,196,.18);
      transform: translateY(-2px);
    }
    .tool-nav-btn.active {
      background: linear-gradient(135deg, #4472c4 0%, #618cef 100%);
      color: #fff;
      border-color: #4472c4;
      pointer-events: none;
      box-shadow: 0 4px 16px rgba(68,114,196,.35);
    }

    /* ===== 使用场景说明区 ===== */
    .usage-card {
      background: #fff;
      border-radius: 10px;
      padding: 16px 20px;
      margin-bottom: 16px;
      box-shadow: 0 1px 4px rgba(0,0,0,.08);
      border-left: 4px solid #4472c4;
    }
    .usage-card .usage-title {
      font-size: 14px;
      font-weight: 600;
      color: #333;
      margin-bottom: 8px;
    }
    .usage-card .usage-text {
      font-size: 13px;
      color: #555;
      line-height: 1.8;
    }

  </style>
</head>
<body>
<div class="container">

  <!-- ════════ 顶部标题 ════════ -->
  <div class="header">
    <div class="header-icon">🔍</div>
    <div>
      <h1>【首页】手车互联PM提效工具</h1>
      <p>输入 Jira 账号和 ID，快速查询当前状态 · 支持钉钉一键批量提醒</p>
    </div>
  </div>

  <!-- ════════ 工具导航 ════════ -->
  <div class="tool-nav">
    <span class="tool-nav-btn active">🔍 Bug 状态查询</span>
    <a class="tool-nav-btn" href="/huawei">🔗 华为链接检查</a>
    <a class="tool-nav-btn" href="/transfer">📋 评论转单分析</a>
  </div>

  <!-- ════════ 使用场景说明 ════════ -->
  <div class="usage-card">
    <div class="usage-title">📝 使用场景说明</div>
    <div class="usage-text">本页面可以批量帮你获取当前jira状态，如果bug所有相关信息，经办人/bug状态等，bug状态为开口的可点击发送按钮钉钉自动帮你群里at对应责任人。</div>
  </div>

  <!-- ════════ 输入区 ════════ -->
  <div class="card">

    <!-- 账号 / 密码 -->
    <div class="form-row">
      <div class="form-group">
        <label>👤 Jira 用户名</label>
        <input type="text"
               class="form-input"
               id="username-input"
               placeholder="GW00295409"
               autocomplete="username">
      </div>
      <div class="form-group">
        <label>🔒 密码</label>
        <div class="pwd-wrap">
          <input type="password"
                 class="form-input"
                 id="password-input"
                 placeholder="输入密码"
                 autocomplete="current-password"
                 style="padding-right: 36px">
          <button class="pwd-toggle"
                  onclick="togglePwd()"
                  id="pwd-toggle-btn">👁</button>
        </div>
      </div>
    </div>

    <!-- 记住密码 -->
    <div class="remember-row">
      <input type="checkbox" id="remember-pwd-checkbox">
      <label for="remember-pwd-checkbox">
        💾 记住密码（下次自动填入）
      </label>
      <span class="clear-saved" id="clear-saved-pwd" onclick="clearSavedPwd()">
        🗑 清除已保存的密码
      </span>
    </div>

    <!-- Jira ID 列表 -->
    <label class="input-label">📋 Jira ID 列表</label>
    <div class="input-hint">
      每行一个，或用空格/逗号分隔，例如：SCHL-1234  SCHL-5678
    </div>
    <textarea id="ids-input"
              placeholder="SCHL-1234&#10;SCHL-5678&#10;SCHL-9999"></textarea>

    <!-- 操作按钮 -->
    <div class="btn-row">
      <button class="btn btn-primary"
              id="query-btn"
              onclick="startQuery()">🚀 开始查询</button>
      <button class="btn btn-secondary"
              onclick="clearAll()">🗑 清空</button>
      <button class="btn btn-export"
              id="export-btn"
              onclick="exportExcel()">📥 导出 Excel</button>
      <button class="btn btn-dingtalk"
              id="dingtalk-btn"
              disabled
              onclick="sendDingTalkNotify()">🔔 一键批量发送</button>
    </div>

    <!-- 进度条 -->
    <div class="progress-wrap" id="progress-wrap">
      <div class="progress-bar-bg">
        <div class="progress-bar-fill" id="progress-fill"></div>
      </div>
      <div class="progress-text" id="progress-text">准备中...</div>
    </div>

    <!-- 日志 -->
    <div class="log-wrap" id="log-wrap"></div>

    <!-- ════════ 钉钉配置区 ════════ -->
    <div class="dingtalk-section">
      <div class="section-title">
        <span>🔔</span> 钉钉机器人配置
      </div>
      <div class="webhook-row">
        <input type="text"
               class="form-input"
               id="webhook-input"
               placeholder="https://oapi.dingtalk.com/robot/send?access_token=xxxxxx">
        <button class="btn-save-webhook" onclick="saveWebhook()">💾 保存</button>
      </div>
      <div class="webhook-hint">
        📌 获取方式：在钉钉群 → 群设置 → 智能群助手 → 添加机器人 → 自定义（webhook）→ 复制链接粘贴到上方<br>
        ✅ 「一键批量发送」将把所有 <strong>未关闭</strong> Bug 按经办人分组，发送到对应钉钉群
      </div>
      <div class="notify-result" id="notify-result"></div>

      <!-- 消息编辑区 -->
      <div class="edit-area-wrap" id="edit-area-wrap">
        <div class="edit-label">✏️ 消息预览与编辑（发送前可修改内容）</div>
        <textarea class="edit-msg-textarea"
                  id="edit-msg-textarea"
                  placeholder="查询后自动生成消息内容，可直接编辑..."></textarea>
        <div class="edit-hint">💡 背景说明和 Bug 列表均可自由修改，点击发送将按编辑后的内容发出，并自动 @对应经办人</div>
      </div>
    </div>
  </div>

  <!-- ════════ 结果区 ════════ -->
  <div class="card" id="result-section">
    <div class="stats-row" id="stats-row"></div>
    <table>
      <thead>
        <tr>
          <th>Jira ID</th>
          <th>状态</th>
          <th>优先级</th>
          <th>经办人</th>
          <th>摘要</th>
        </tr>
      </thead>
      <tbody id="result-tbody">
        <tr>
          <td colspan="5" class="empty-state">暂无数据</td>
        </tr>
      </tbody>
    </table>
  </div>

</div>

<script>
  /* ═══════════════════════════════════════
     全局状态
     ═══════════════════════════════════════ */
  let allResults = [];
  let isQuerying  = false;

  /* ═══════════════════════════════════════
     页面初始化：恢复记住的用户名 & 密码 & Webhook
     ═══════════════════════════════════════ */
  (function() {
    const savedUser    = localStorage.getItem("jira_username");
    const savedPwd     = localStorage.getItem("jira_password");
    const savedWebhook = localStorage.getItem("dingtalk_webhook");
    if (savedUser)    document.getElementById("username-input").value = savedUser;
    if (savedWebhook) document.getElementById("webhook-input").value  = savedWebhook;

    // 恢复密码
    if (savedPwd) {
      document.getElementById("password-input").value = savedPwd;
      document.getElementById("remember-pwd-checkbox").checked = true;
      document.getElementById("clear-saved-pwd").classList.add("visible");
    }

    // 也尝试从服务端读取 webhook（config.json）
    fetch("/api/config")
      .then(r => r.json())
      .then(data => {
        if (data.dingtalk_webhook && !savedWebhook) {
          document.getElementById("webhook-input").value = data.dingtalk_webhook;
        }
      })
      .catch(() => {});
  })();

  /* ═══════════════════════════════════════
     密码 显示/隐藏 切换
     ═══════════════════════════════════════ */
  function togglePwd() {
    const inp = document.getElementById("password-input");
    const btn = document.getElementById("pwd-toggle-btn");
    if (inp.type === "password") {
      inp.type = "text";
      btn.textContent = "🙈";
    } else {
      inp.type = "password";
      btn.textContent = "👁";
    }
  }

  /* ═══════════════════════════════════════
     工具函数
     ═══════════════════════════════════════ */

  // 解析输入框文本 → ID 数组（支持换行/空格/逗号分隔）
  function parseIds(raw) {
    return raw
      .split(/[\n,\s]+/)
      .map(s => s.trim().toUpperCase())
      .filter(s => s.length > 0);
  }

  // 向日志面板追加一行
  function log(msg, type) {
    type = type || "info";
    const wrap = document.getElementById("log-wrap");
    wrap.style.display = "block";
    const line = document.createElement("div");
    line.className = "log-line log-" + type;
    line.textContent = msg;
    wrap.appendChild(line);
    wrap.scrollTop = wrap.scrollHeight;
  }

  // 更新进度条
  function setProgress(current, total) {
    const pct = total > 0 ? Math.round(current / total * 100) : 0;
    document.getElementById("progress-fill").style.width = pct + "%";
    document.getElementById("progress-text").textContent =
      "进度：" + current + " / " + total + "  (" + pct + "%)";
  }

  /* ═══════════════════════════════════════
     渲染：状态徽章
     ═══════════════════════════════════════ */
  function badgeHtml(status, color, isError) {
    if (isError) {
      return '<span class="badge badge-error">❌ ' + status + '</span>';
    }
    const cls = "badge-" + (color || "gray");
    return '<span class="badge ' + cls + '">' + status + '</span>';
  }

  /* ═══════════════════════════════════════
     渲染：优先级图标
     ═══════════════════════════════════════ */
  const PRIORITY_MAP = {
    "最高": "🔴", "High": "🔴", "P0": "🔴",
    "高":   "🟠", "P1": "🟠",
    "中":   "🟡", "Medium": "🟡", "P2": "🟡",
    "低":   "🟢", "Low": "🟢", "P3": "🟢",
    "最低": "⚪", "Lowest": "⚪"
  };

  function priorityIcon(p) {
    return (PRIORITY_MAP[p] || "") + " " + p;
  }

  /* ═══════════════════════════════════════
     渲染：结果表格
     ═══════════════════════════════════════ */
  function renderTable(results) {
    const tbody = document.getElementById("result-tbody");

    if (results.length === 0) {
      tbody.innerHTML =
        '<tr><td colspan="5" class="empty-state">暂无数据</td></tr>';
      return;
    }

    const baseUrl = "https://jira.gwm.cn";

    tbody.innerHTML = results.map(function(r) {
      // 错误行
      if (r.error) {
        return '<tr>'
          + '<td><span class="id-link">' + r.id + '</span></td>'
          + '<td>' + badgeHtml(r.error, "error", true) + '</td>'
          + '<td>—</td><td>—</td><td>—</td>'
          + '</tr>';
      }

      // 正常行
      const idLink =
        '<a class="id-link" href="' + baseUrl + '/browse/' + r.id
        + '" target="_blank">' + r.id + '</a>';
      const summaryEsc =
        r.summary.replace(/</g, "&lt;").replace(/>/g, "&gt;");

      return '<tr>'
        + '<td>' + idLink + '</td>'
        + '<td>' + badgeHtml(r.status, r.status_color, false) + '</td>'
        + '<td style="white-space:nowrap;font-size:12px">'
          + priorityIcon(r.priority) + '</td>'
        + '<td style="white-space:nowrap;font-size:12px;color:#666">'
          + r.assignee + '</td>'
        + '<td class="summary-cell" title="' + summaryEsc + '">'
          + summaryEsc + '</td>'
        + '</tr>';
    }).join("");
  }

  /* ═══════════════════════════════════════
     渲染：统计标签
     ═══════════════════════════════════════ */
  const DONE_STATUSES = ["完成", "关闭", "已关闭", "已解决", "Done", "Closed", "Resolved"];

  function renderStats(results) {
    const total  = results.length;
    const done   = results.filter(r => !r.error && DONE_STATUSES.includes(r.status)).length;
    const open   = results.filter(r => !r.error && !DONE_STATUSES.includes(r.status)).length;
    const err    = results.filter(r => r.error).length;

    let html = '<div class="stat-chip">'
      + '<span class="dot dot-total"></span>共 ' + total + ' 个</div>'
      + '<div class="stat-chip">'
      + '<span class="dot dot-done"></span>已完成 ' + done + ' 个</div>'
      + '<div class="stat-chip">'
      + '<span class="dot dot-open"></span>未关闭 ' + open + ' 个</div>';

    if (err > 0) {
      html += '<div class="stat-chip">'
        + '<span class="dot dot-error"></span>查询失败 ' + err + ' 个</div>';
    }
    if (open > 0) {
      html += '<div class="stat-chip" style="background:#e8f9e8">'
        + '<span class="dot dot-notify"></span>'
        + '<span style="color:#18a058">可发钉钉提醒</span></div>';
    }

    document.getElementById("stats-row").innerHTML = html;
  }

  /* ═══════════════════════════════════════
     核心：主查询逻辑
     ═══════════════════════════════════════ */
  async function startQuery() {
    if (isQuerying) return;

    // 收集输入
    const raw      = document.getElementById("ids-input").value.trim();
    const ids      = parseIds(raw);
    const username = document.getElementById("username-input").value.trim();
    const password = document.getElementById("password-input").value;

    // 校验
    if (ids.length === 0) { alert("请先输入 Jira ID"); return; }
    if (!username || !password) { alert("请输入 Jira 用户名和密码"); return; }

    // 记住用户名
    localStorage.setItem("jira_username", username);

    // 记住/清除密码
    const rememberPwd = document.getElementById("remember-pwd-checkbox").checked;
    const clearBtn    = document.getElementById("clear-saved-pwd");
    if (rememberPwd) {
      localStorage.setItem("jira_password", password);
      clearBtn.classList.add("visible");
    } else {
      localStorage.removeItem("jira_password");
      clearBtn.classList.remove("visible");
    }

    // 重置 UI
    isQuerying = true;
    allResults = [];
    document.getElementById("query-btn").disabled          = true;
    document.getElementById("export-btn").style.display    = "none";
    document.getElementById("dingtalk-btn").disabled        = true;
    document.getElementById("progress-wrap").style.display = "block";
    document.getElementById("log-wrap").innerHTML          = "";
    document.getElementById("log-wrap").style.display      = "block";
    document.getElementById("result-section").style.display= "none";
    document.getElementById("notify-result").style.display = "none";
    document.getElementById("edit-area-wrap").style.display = "none";
    setProgress(0, ids.length);

    log("开始查询，共 " + ids.length + " 个 ID...", "info");

    // 逐个查询
    for (let i = 0; i < ids.length; i++) {
      const id = ids[i];
      log("[" + (i + 1) + "/" + ids.length + "] 查询 " + id + " ...", "info");

      try {
        const resp = await fetch("/api/status", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ issue_id: id, username: username, password: password })
        });
        const data = await resp.json();
        allResults.push(data);

        if (data.error) {
          log("  ❌ " + id + "：" + data.error, "err");
        } else {
          log("  ✅ " + id + "：" + data.status + " | " + data.assignee, "ok");
        }
      } catch (e) {
        allResults.push({ id: id, error: "请求失败：" + e.message });
        log("  ❌ " + id + "：请求失败", "err");
      }

      setProgress(i + 1, ids.length);
    }

    // 统计未关闭（排除法：不在 DONE_STATUSES 的都算未关闭）
    const openCount = allResults.filter(
      r => !r.error && !DONE_STATUSES.includes(r.status)
    ).length;

    // 展示结果
    renderTable(allResults);
    renderStats(allResults);
    document.getElementById("result-section").style.display = "block";
    document.getElementById("export-btn").style.display     = "inline-block";

    // 有未关闭 Bug 才启用钉钉按钮 + 显示编辑区
    const dingBtn = document.getElementById("dingtalk-btn");
    const editWrap = document.getElementById("edit-area-wrap");
    const editTextarea = document.getElementById("edit-msg-textarea");
    if (openCount > 0) {
      dingBtn.disabled = false;
      log("🔔 检测到 " + openCount + " 个未关闭 Bug，可编辑消息后「一键批量发送」", "warn");

      // 请求服务端生成消息预览
      try {
        const previewResp = await fetch("/api/preview-msg", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ results: allResults })
        });
        const previewData = await previewResp.json();
        if (previewData.success && previewData.text) {
          editTextarea.value = previewData.text;
        } else {
          editTextarea.value = "（生成预览失败：" + (previewData.error || "未知错误") + "）";
        }
      } catch (e) {
        editTextarea.value = "（生成预览失败：" + e.message + "）";
      }
      editWrap.style.display = "block";
    } else {
      dingBtn.disabled = true;
      editWrap.style.display = "none";
    }

    document.getElementById("query-btn").disabled = false;
    isQuerying = false;
    log("查询完成 ✅", "ok");
    document.getElementById("result-section").scrollIntoView({ behavior: "smooth" });
  }

  /* ═══════════════════════════════════════
     钉钉：保存 Webhook
     ═══════════════════════════════════════ */
  async function saveWebhook() {
    const url = document.getElementById("webhook-input").value.trim();
    if (!url) { alert("请先填写 Webhook URL"); return; }
    if (!url.startsWith("https://oapi.dingtalk.com/")) {
      alert("Webhook 格式不对，应以 https://oapi.dingtalk.com/ 开头");
      return;
    }

    // 本地存储
    localStorage.setItem("dingtalk_webhook", url);

    // 发送到服务端保存到 config.json
    try {
      const resp = await fetch("/api/save-webhook", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ webhook: url })
      });
      const data = await resp.json();
      if (data.success) {
        alert("✅ Webhook 已保存（本地 + config.json）");
      } else {
        alert("本地已保存，但写入 config.json 失败：" + data.error);
      }
    } catch (e) {
      alert("本地已保存，服务端保存失败：" + e.message);
    }
  }

  /* ═══════════════════════════════════════
     钉钉：一键批量发送
     ═══════════════════════════════════════ */
  async function sendDingTalkNotify() {
    if (allResults.length === 0) {
      alert("请先查询 Jira 状态");
      return;
    }

    const webhookUrl = document.getElementById("webhook-input").value.trim();
    if (!webhookUrl) {
      alert("请先在下方填写钉钉机器人 Webhook URL");
      document.getElementById("webhook-input").focus();
      return;
    }

    const btn = document.getElementById("dingtalk-btn");
    btn.disabled    = true;
    btn.textContent = "⏳ 发送中...";

    const notifyDiv = document.getElementById("notify-result");
    notifyDiv.style.display = "none";

    try {
      const resp = await fetch("/api/notify", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({
          results: allResults,
          webhook: webhookUrl,
          custom_text: document.getElementById("edit-msg-textarea").value.trim()
        })
      });
      const data = await resp.json();

      notifyDiv.style.display = "block";
      if (data.success) {
        notifyDiv.className = "notify-result success";
        notifyDiv.innerHTML =
          "✅ 发送成功！共提醒 <strong>" + data.open_count + "</strong> 个未关闭 Bug，"
          + "涉及 <strong>" + data.assignee_count + "</strong> 位经办人：<br>"
          + data.assignees.map(a => "👤 " + a).join("　");
      } else {
        notifyDiv.className = "notify-result error";
        notifyDiv.innerHTML = "❌ 发送失败：" + data.error;
      }
    } catch (e) {
      notifyDiv.style.display = "block";
      notifyDiv.className     = "notify-result error";
      notifyDiv.innerHTML     = "❌ 请求失败：" + e.message;
    }

    btn.disabled    = false;
    btn.textContent = "🔔 一键批量发送到钉钉群";
    notifyDiv.scrollIntoView({ behavior: "smooth", block: "nearest" });
  }

  /* ═══════════════════════════════════════
     操作：清除已保存的密码
     ═══════════════════════════════════════ */
  function clearSavedPwd() {
    localStorage.removeItem("jira_password");
    document.getElementById("password-input").value      = "";
    document.getElementById("remember-pwd-checkbox").checked = false;
    document.getElementById("clear-saved-pwd").classList.remove("visible");
  }

  /* ═══════════════════════════════════════
     操作：清空
     ═══════════════════════════════════════ */
  function clearAll() {
    document.getElementById("ids-input").value          = "";
    document.getElementById("password-input").value      = "";
    document.getElementById("log-wrap").innerHTML        = "";
    document.getElementById("log-wrap").style.display    = "none";
    document.getElementById("progress-wrap").style.display = "none";
    document.getElementById("result-section").style.display = "none";
    document.getElementById("export-btn").style.display  = "none";
    document.getElementById("dingtalk-btn").disabled       = true;
    document.getElementById("notify-result").style.display = "none";
    document.getElementById("edit-area-wrap").style.display = "none";
    allResults = [];
  }

  /* ═══════════════════════════════════════
     操作：导出 Excel
     ═══════════════════════════════════════ */
  async function exportExcel() {
    if (allResults.length === 0) return;

    const btn = document.getElementById("export-btn");
    btn.disabled    = true;
    btn.textContent = "⏳ 导出中...";
    try {
      const resp = await fetch("/api/export", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ results: allResults })
      });
      const data = await resp.json();
      if (data.success) {
        alert("导出成功！\n文件保存至：" + data.path);
      } else {
        alert("导出失败：" + data.error);
      }
    } catch (e) {
      alert("导出请求失败：" + e.message);
    }
    btn.disabled    = false;
    btn.textContent = "📥 导出 Excel";
  }

  /* ═══════════════════════════════════════
     快捷键：Ctrl+Enter 触发查询
     ═══════════════════════════════════════ */
  document.addEventListener("keydown", function(e) {
    if (e.ctrlKey && e.key === "Enter") { startQuery(); }
  });

</script>
</body>
</html>"""


# ─────────────────────────────────────────────
# 华为链接检查 —— HTML 页面
# ─────────────────────────────────────────────
HUAWEI_HTML_PAGE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Jira Huawei 链接检查工具</title>
  <style>
    /* ===== 全局重置 ===== */
    * {
      margin: 0;
      padding: 0;
      box-sizing: border-box;
    }

    body {
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      background: #f0f2f5;
      color: #1a1a2e;
      min-height: 100vh;
    }

    /* ===== 布局容器 ===== */
    .container {
      max-width: 1100px;
      margin: 0 auto;
      padding: 20px;
    }

    h1 {
      font-size: 22px;
      font-weight: 600;
      text-align: center;
      padding: 20px 0 6px;
      color: #16213e;
    }

    .subtitle {
      text-align: center;
      color: #8c8c8c;
      font-size: 13px;
      margin-bottom: 24px;
    }

    /* ===== 卡片 ===== */
    .card {
      background: #fff;
      border-radius: 10px;
      padding: 20px 24px;
      margin-bottom: 16px;
      box-shadow: 0 1px 4px rgba(0, 0, 0, .08);
    }

    .card h2 {
      font-size: 15px;
      font-weight: 600;
      color: #16213e;
      margin-bottom: 16px;
      padding-bottom: 10px;
      border-bottom: 1px solid #eee;
    }

    /* ===== 表单 ===== */
    .form-row {
      display: flex;
      gap: 12px;
      margin-bottom: 12px;
      flex-wrap: wrap;
    }

    .form-group {
      flex: 1;
      min-width: 180px;
    }

    .form-group label {
      display: block;
      font-size: 13px;
      color: #555;
      margin-bottom: 5px;
      font-weight: 500;
    }

    .form-group input,
    .form-group textarea {
      width: 100%;
      padding: 8px 12px;
      border: 1px solid #d9d9d9;
      border-radius: 6px;
      font-size: 13px;
      transition: border-color .2s;
    }

    .form-group input:focus,
    .form-group textarea:focus {
      outline: none;
      border-color: #4472c4;
      box-shadow: 0 0 0 2px rgba(68, 114, 196, .15);
    }

    .form-group textarea {
      min-height: 100px;
      resize: vertical;
      font-family: monospace;
      font-size: 12px;
    }

    .hint {
      font-size: 11px;
      color: #aaa;
      margin-top: 3px;
    }

    /* ===== 按钮 ===== */
    .btn-row {
      display: flex;
      gap: 10px;
      align-items: center;
      margin-top: 16px;
      flex-wrap: wrap;
    }

    .btn {
      padding: 8px 20px;
      border: none;
      border-radius: 6px;
      font-size: 13px;
      font-weight: 500;
      cursor: pointer;
      transition: all .2s;
      display: inline-flex;
      align-items: center;
      gap: 6px;
    }

    .btn-primary             { background: #4472c4; color: #fff; }
    .btn-primary:hover       { background: #355da0; }
    .btn-primary:disabled    { background: #a0b8d8; cursor: not-allowed; }

    .btn-secondary           { background: #f0f0f0; color: #333; }
    .btn-secondary:hover     { background: #e0e0e0; }

    .btn-export              { background: #28a745; color: #fff; }
    .btn-export:hover        { background: #218838; }
    .btn-export:disabled     { background: #8fbc9a; cursor: not-allowed; }

    /* ===== 进度条 ===== */
    .progress-bar {
      height: 3px;
      background: #e8e8e8;
      border-radius: 2px;
      margin-top: 14px;
      overflow: hidden;
      display: none;
    }

    .progress-bar .fill {
      height: 100%;
      background: #4472c4;
      border-radius: 2px;
      transition: width .3s;
      width: 0;
    }

    /* ===== 日志 ===== */
    .log {
      margin-top: 12px;
      max-height: 120px;
      overflow-y: auto;
      font-size: 12px;
      color: #666;
      background: #fafafa;
      padding: 8px 12px;
      border-radius: 6px;
      font-family: monospace;
      line-height: 1.8;
      display: none;
    }

    .log.show  { display: block; }
    .log .ok   { color: #28a745; }
    .log .err  { color: #dc3545; }
    .log .info { color: #4472c4; }

    /* ===== 统计卡片 ===== */
    .stats {
      display: flex;
      gap: 16px;
      flex-wrap: wrap;
      margin: 12px 0;
    }

    .stat-item {
      background: #f7f8fa;
      border-radius: 8px;
      padding: 12px 18px;
      min-width: 120px;
      text-align: center;
    }

    .stat-item .num {
      font-size: 24px;
      font-weight: 700;
      color: #16213e;
    }

    .stat-item .label {
      font-size: 12px;
      color: #888;
      margin-top: 2px;
    }

    /* ===== 表格 ===== */
    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 12px;
      margin-top: 12px;
    }

    thead th {
      background: #4472c4;
      color: #fff;
      padding: 8px 10px;
      text-align: left;
      font-weight: 500;
      position: sticky;
      top: 0;
      white-space: nowrap;
    }

    tbody td {
      padding: 7px 10px;
      border-bottom: 1px solid #f0f0f0;
      vertical-align: top;
      max-width: 300px;
      word-break: break-all;
    }

    tbody tr:hover { background: #f8f9ff; }

    /* ===== 标签 ===== */
    .tag {
      display: inline-block;
      padding: 2px 8px;
      border-radius: 10px;
      font-size: 11px;
      font-weight: 500;
    }

    .tag-direct { background: #e6f7ee; color: #1a8a4a; }
    .tag-linked { background: #fff3e0; color: #c77d00; }

    /* ===== 其他 ===== */
    .link-cell {
      max-width: 200px;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }

    .snippet {
      color: #666;
      font-size: 11px;
      max-width: 250px;
      display: block;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }

    .no-data {
      text-align: center;
      padding: 30px;
      color: #aaa;
      font-size: 14px;
    }

    /* ===== Toast 提示 ===== */
    .toast {
      position: fixed;
      top: 20px;
      right: 20px;
      padding: 10px 20px;
      border-radius: 8px;
      color: #fff;
      font-size: 13px;
      z-index: 999;
      opacity: 0;
      transition: opacity .3s;
      pointer-events: none;
    }

    .toast.show       { opacity: 1; }
    .toast-success    { background: #28a745; }
    .toast-error      { background: #dc3545; }

    /* ===== 复选框行 ===== */
    .checkbox-row {
      display: flex;
      align-items: center;
      gap: 6px;
      margin-top: 8px;
    }

    .checkbox-row label {
      font-size: 12px;
      color: #666;
      cursor: pointer;
    }

    /* ===== 工具导航栏 ===== */
    .tool-nav {
      display: flex;
      gap: 14px;
      margin-bottom: 20px;
      flex-wrap: wrap;
      justify-content: center;
      padding: 14px 20px;
      background: #fff;
      border-radius: 14px;
      box-shadow: 0 2px 12px rgba(0,0,0,.08);
    }
    .tool-nav-btn {
      padding: 11px 28px;
      border: 2px solid transparent;
      border-radius: 12px;
      font-size: 15px;
      font-weight: 600;
      text-decoration: none;
      cursor: pointer;
      transition: all .25s;
      display: inline-flex;
      align-items: center;
      gap: 8px;
      background: #f2f4f8;
      color: #4a5568;
      box-shadow: 0 1px 3px rgba(0,0,0,.05);
    }
    .tool-nav-btn:hover {
      border-color: #4472c4;
      color: #4472c4;
      background: #eef3ff;
      box-shadow: 0 3px 12px rgba(68,114,196,.18);
      transform: translateY(-2px);
    }
    .tool-nav-btn.active {
      background: linear-gradient(135deg, #4472c4 0%, #618cef 100%);
      color: #fff;
      border-color: #4472c4;
      pointer-events: none;
      box-shadow: 0 4px 16px rgba(68,114,196,.35);
    }

    /* ===== 使用场景说明区 ===== */
    .usage-card {
      background: #fff;
      border-radius: 10px;
      padding: 16px 20px;
      margin-bottom: 16px;
      box-shadow: 0 1px 4px rgba(0,0,0,.08);
      border-left: 4px solid #4472c4;
    }
    .usage-card .usage-title {
      font-size: 14px;
      font-weight: 600;
      color: #333;
      margin-bottom: 8px;
    }
    .usage-card .usage-text {
      font-size: 13px;
      color: #555;
      line-height: 1.8;
    }
  </style>
</head>
<body>
  <div class="container">

    <h1>Jira Huawei 链接检查工具</h1>
    <p class="subtitle">输入账号密码和 Bug ID，自动检查备注中的 Huawei 链接</p>

    <!-- 工具导航 -->
    <div class="tool-nav">
      <a class="tool-nav-btn" href="/">🔍 Bug 状态查询</a>
      <span class="tool-nav-btn active">🔗 华为链接检查</span>
      <a class="tool-nav-btn" href="/transfer">📋 评论转单分析</a>
    </div>

    <!-- 使用场景说明 -->
    <div class="usage-card">
      <div class="usage-title">📝 使用场景说明</div>
      <div class="usage-text">本页面主要在实际工作中，需要批量催手机侧的bug，需要找到cpm单号，以往需人肉点击jira==》查看备注==》找到cpm，此页面可替你批量导出CPM（同时支持关联的子票分析，目前只支持一级），你批量发送给手机厂商追bug即可</div>
    </div>

    <!-- 配置卡片 -->
    <div class="card">
      <h2>配置信息</h2>

      <!-- Jira 地址 -->
      <div class="form-row">
        <div class="form-group" style="max-width: 320px">
          <label>Jira 服务地址</label>
          <input type="text" id="jiraUrl" value="https://jira.gwm.cn" placeholder="https://jira.gwm.cn">
        </div>
      </div>

      <!-- 用户名 / 密码 -->
      <div class="form-row">
        <div class="form-group">
          <label>用户名</label>
          <input type="text" id="username" placeholder="GW00295409" value="GW00295409">
        </div>
        <div class="form-group">
          <label>密码</label>
          <input type="password" id="password" placeholder="Jira 密码" value="5352Zhao123456.">
        </div>
      </div>

      <!-- Bug ID 列表 -->
      <div class="form-row">
        <div class="form-group">
          <label>Bug ID 列表</label>
          <textarea id="bugList" placeholder="SCHL-5918&#10;SCHL-3731&#10;SCHL-5207&#10;支持逗号、空格、换行分隔"></textarea>
          <div class="hint">多个 ID 用逗号、空格或换行分隔</div>
        </div>
      </div>

      <!-- 关联检查开关 -->
      <div class="checkbox-row">
        <input type="checkbox" id="checkLinked" checked>
        <label for="checkLinked">启用关联 Issue 检查（自身未命中时，自动检查关联 Issue 的备注）</label>
      </div>

      <!-- 操作按钮 -->
      <div class="btn-row">
        <button class="btn btn-primary" id="btnStart" onclick="startAnalysis()">开始分析</button>
        <button class="btn btn-secondary" onclick="document.getElementById('log').classList.toggle('show')">显示/隐藏日志</button>
      </div>

      <!-- 进度条 -->
      <div class="progress-bar" id="progressBar">
        <div class="fill" id="progressFill"></div>
      </div>

      <!-- 日志区 -->
      <div class="log" id="log"></div>
    </div>

    <!-- 结果卡片（默认隐藏） -->
    <div id="resultCard" class="card" style="display: none">
      <h2>分析结果</h2>
      <div class="stats" id="stats"></div>
      <div class="btn-row">
        <button class="btn btn-export" id="btnExport" onclick="exportExcel()">导出 Excel (.xlsx)</button>
        <span id="resultTime" style="font-size: 12px; color: #aaa"></span>
      </div>
      <div style="overflow-x: auto; max-height: 500px; overflow-y: auto; margin-top: 10px" id="tableWrap"></div>
    </div>

  </div><!-- /.container -->

  <!-- Toast 提示 -->
  <div class="toast" id="toast"></div>

  <script>
    let allResults = [];

    /* ---------- 工具函数 ---------- */

    function log(msg, type) {
      const el = document.getElementById('log');
      el.classList.add('show');
      el.innerHTML += '<div class="' + type + '">' + msg + '</div>';
      el.scrollTop = el.scrollHeight;
    }

    function showToast(msg, type) {
      const t = document.getElementById('toast');
      t.textContent = msg;
      t.className = 'toast toast-' + type + ' show';
      setTimeout(() => t.className = 'toast', 3000);
    }

    function setProgress(pct) {
      const bar  = document.getElementById('progressBar');
      const fill = document.getElementById('progressFill');
      bar.style.display = pct > 0 ? 'block' : 'none';
      fill.style.width  = pct + '%';
    }

    function escHtml(s) {
      if (!s) return '';
      return s
        .replace(/&/g,  '&amp;')
        .replace(/</g,  '&lt;')
        .replace(/>/g,  '&gt;')
        .replace(/"/g,  '&quot;');
    }

    /* ---------- 主逻辑：开始分析 ---------- */

    async function startAnalysis() {
      const url         = document.getElementById('jiraUrl').value.trim().replace(/\/+$/, '');
      const username    = document.getElementById('username').value.trim();
      const password    = document.getElementById('password').value.trim();
      const bugText     = document.getElementById('bugList').value.trim();
      const checkLinked = document.getElementById('checkLinked').checked;

      // 校验
      if (!url || !username || !password) {
        showToast('请填写完整的配置信息', 'error');
        return;
      }
      if (!bugText) {
        showToast('请输入至少一个 Bug ID', 'error');
        return;
      }

      const ids = bugText.split(/[,，\s\n]+/).map(s => s.trim()).filter(Boolean);
      if (!ids.length) {
        showToast('未检测到有效 Bug ID', 'error');
        return;
      }

      // 重置界面
      const btn = document.getElementById('btnStart');
      btn.disabled    = true;
      btn.textContent = '分析中...';
      document.getElementById('resultCard').style.display = 'none';
      document.getElementById('log').innerHTML = '';
      allResults = [];
      setProgress(0);

      log('开始分析 ' + ids.length + ' 个 Issue...', 'info');

      // 逐条请求
      for (let i = 0; i < ids.length; i++) {
        const id  = ids[i];
        const pct = Math.round((i / ids.length) * 100);
        setProgress(pct);
        log('--- [' + id + '] (' + (i + 1) + '/' + ids.length + ') ---', 'info');

        try {
          const resp = await fetch('/api/huawei/check', {
            method:  'POST',
            headers: { 'Content-Type': 'application/json' },
            body:    JSON.stringify({
              jira_url:     url,
              username:     username,
              password:     password,
              issue_id:     id,
              check_linked: checkLinked
            })
          });
          const data = await resp.json();

          if (data.error) {
            log('  ' + data.error, 'err');
          } else if (data.hits && data.hits.length > 0) {
            log('  发现 ' + data.hits.length + ' 条 Huawei 链接', 'ok');
            data.hits.forEach(h => {
              log('    备注 #' + h.comment_id + ' | ' + h.author + ' | ' + h.huawei_links.substring(0, 80) + '...', 'ok');
            });
            allResults = allResults.concat(data.hits);
          } else {
            log('  未发现 Huawei 链接', 'info');
          }
        } catch (e) {
          log('  请求失败: ' + e.message, 'err');
        }
      }

      // 完成
      setProgress(100);
      btn.disabled    = false;
      btn.textContent = '开始分析';
      log('分析完成！共命中 ' + allResults.length + ' 条', allResults.length > 0 ? 'ok' : 'info');

      if (allResults.length > 0) {
        renderResults();
        showToast('分析完成，发现 ' + allResults.length + ' 条结果', 'success');
      } else {
        showToast('分析完成，未发现 Huawei 链接', 'error');
      }
    }

    /* ---------- 渲染结果表格 ---------- */

    function renderResults() {
      const card = document.getElementById('resultCard');
      card.style.display = 'block';

      const direct = allResults.filter(r => r.source === '直接命中').length;
      const linked = allResults.filter(r => r.source !== '直接命中').length;

      // 统计栏
      document.getElementById('stats').innerHTML =
        '<div class="stat-item">' +
          '<div class="num">' + allResults.length + '</div>' +
          '<div class="label">命中总数</div>' +
        '</div>' +
        '<div class="stat-item">' +
          '<div class="num" style="color:#1a8a4a">' + direct + '</div>' +
          '<div class="label">直接命中</div>' +
        '</div>' +
        '<div class="stat-item">' +
          '<div class="num" style="color:#c77d00">' + linked + '</div>' +
          '<div class="label">关联Issue命中</div>' +
        '</div>';

      // 数据表格
      let html = '<table>' +
        '<thead><tr>' +
          '<th>Issue ID</th><th>Jira链接</th><th>来源</th><th>父Issue</th>' +
          '<th>备注ID</th><th>作者</th><th>时间</th><th>Huawei链接</th><th>备注摘要</th>' +
        '</tr></thead>' +
        '<tbody>';

      allResults.forEach(r => {
        const tagClass = r.source === '直接命中' ? 'tag-direct' : 'tag-linked';
        html += '<tr>' +
          '<td><b>' + r.issue_id + '</b></td>' +
          '<td><a href="' + r.jira_url + '" target="_blank" style="color:#4472c4;text-decoration:none">' + r.issue_id + '</a></td>' +
          '<td><span class="tag ' + tagClass + '">' + r.source + '</span></td>' +
          '<td>' + (r.parent_issue_id || '-') + '</td>' +
          '<td>' + r.comment_id + '</td>' +
          '<td>' + r.author + '</td>' +
          '<td style="white-space:nowrap">' + r.created + '</td>' +
          '<td class="link-cell" title="' + escHtml(r.huawei_links) + '">' + escHtml(r.huawei_links.substring(0, 60)) + '...</td>' +
          '<td><span class="snippet" title="' + escHtml(r.comment_snippet) + '">' + escHtml(r.comment_snippet) + '</span></td>' +
        '</tr>';
      });

      html += '</tbody></table>';
      document.getElementById('tableWrap').innerHTML = html;
      document.getElementById('resultTime').textContent = '分析时间: ' + new Date().toLocaleString('zh-CN');
    }

    /* ---------- 导出 Excel ---------- */

    function exportExcel() {
      if (!allResults.length) return;
      const data  = JSON.stringify(allResults);
      const form  = document.createElement('form');
      form.method = 'POST';
      form.action = '/api/huawei/export';
      const input   = document.createElement('input');
      input.type    = 'hidden';
      input.name    = 'data';
      input.value   = data;
      form.appendChild(input);
      document.body.appendChild(form);
      form.submit();
      document.body.removeChild(form);
    }
  </script>
</body>
</html>"""


# ─────────────────────────────────────────────
# 评论转单分析 —— HTML 页面
# ─────────────────────────────────────────────
TRANSFER_HTML_PAGE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Jira 评论转单分析工具</title>
  <style>
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body {
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      background: #f0f2f5; color: #1a1a2e; min-height: 100vh;
    }
    .container { max-width: 1100px; margin: 0 auto; padding: 20px; }
    h1 { font-size: 22px; font-weight: 600; text-align: center; padding: 20px 0 6px; color: #16213e; }
    .subtitle { text-align: center; color: #8c8c8c; font-size: 13px; margin-bottom: 24px; }
    .card {
      background: #fff; border-radius: 10px; padding: 20px 24px; margin-bottom: 16px;
      box-shadow: 0 1px 4px rgba(0,0,0,.08);
    }
    .card h2 { font-size: 15px; font-weight: 600; color: #16213e; margin-bottom: 16px; padding-bottom: 10px; border-bottom: 1px solid #eee; }
    .form-row { display: flex; gap: 12px; margin-bottom: 12px; flex-wrap: wrap; }
    .form-group { flex: 1; min-width: 180px; }
    .form-group label { display: block; font-size: 13px; color: #555; margin-bottom: 5px; font-weight: 500; }
    .form-group input, .form-group textarea {
      width: 100%; padding: 8px 12px; border: 1px solid #d9d9d9; border-radius: 6px;
      font-size: 13px; transition: border-color .2s;
    }
    .form-group input:focus, .form-group textarea:focus {
      outline: none; border-color: #4472c4; box-shadow: 0 0 0 2px rgba(68,114,196,.15);
    }
    .form-group textarea { min-height: 120px; resize: vertical; font-family: monospace; font-size: 12px; }
    .hint { font-size: 11px; color: #aaa; margin-top: 3px; }
    .btn-row { display: flex; gap: 10px; align-items: center; margin-top: 16px; flex-wrap: wrap; }
    .btn {
      padding: 8px 20px; border: none; border-radius: 6px; font-size: 13px;
      font-weight: 500; cursor: pointer; transition: all .2s; display: inline-flex; align-items: center; gap: 6px;
    }
    .btn-primary { background: #4472c4; color: #fff; }
    .btn-primary:hover { background: #355da0; }
    .btn-primary:disabled { background: #a0b8d8; cursor: not-allowed; }
    .btn-secondary { background: #f0f0f0; color: #333; }
    .btn-secondary:hover { background: #e0e0e0; }
    .btn-export { background: #28a745; color: #fff; }
    .btn-export:hover { background: #218838; }
    .btn-export:disabled { background: #8fbc9a; cursor: not-allowed; }
    .progress-bar { height: 3px; background: #e8e8e8; border-radius: 2px; margin-top: 14px; overflow: hidden; display: none; }
    .progress-bar .fill { height: 100%; background: #4472c4; border-radius: 2px; transition: width .3s; width: 0; }
    .log { margin-top: 12px; max-height: 150px; overflow-y: auto; font-size: 12px; color: #666; background: #fafafa; padding: 8px 12px; border-radius: 6px; font-family: monospace; line-height: 1.8; display: none; }
    .log.show { display: block; }
    .log .ok { color: #28a745; }
    .log .err { color: #dc3545; }
    .log .info { color: #4472c4; }
    .stats { display: flex; gap: 16px; flex-wrap: wrap; margin: 12px 0; }
    .stat-item { background: #f7f8fa; border-radius: 8px; padding: 12px 18px; min-width: 100px; text-align: center; }
    .stat-item .num { font-size: 24px; font-weight: 700; color: #16213e; }
    .stat-item .label { font-size: 12px; color: #888; margin-top: 2px; }
    table { width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 12px; }
    thead th { background: #4472c4; color: #fff; padding: 8px 10px; text-align: left; font-weight: 500; position: sticky; top: 0; white-space: nowrap; }
    tbody td { padding: 7px 10px; border-bottom: 1px solid #f0f0f0; vertical-align: top; max-width: 280px; word-break: break-all; }
    tbody tr:hover { background: #f8f9ff; }
    .author-card { background: #f7f8ff; border-radius: 8px; padding: 10px 14px; min-width: 80px; text-align: center; }
    .author-card .name { font-size: 13px; font-weight: 600; color: #1a1a2e; }
    .author-card .count { font-size: 20px; font-weight: 700; color: #4472c4; }
    .no-data { text-align: center; padding: 30px; color: #aaa; font-size: 14px; }
    .toast { position: fixed; top: 20px; right: 20px; padding: 10px 20px; border-radius: 8px; color: #fff; font-size: 13px; z-index: 999; opacity: 0; transition: opacity .3s; pointer-events: none; }
    .toast.show { opacity: 1; }
    .toast-success { background: #28a745; }
    .toast-error { background: #dc3545; }
    .tool-nav {
      display: flex; gap: 14px; margin-bottom: 20px; flex-wrap: wrap;
      justify-content: center; padding: 14px 20px; background: #fff;
      border-radius: 14px; box-shadow: 0 2px 12px rgba(0,0,0,.08);
    }
    .tool-nav-btn {
      padding: 11px 28px; border: 2px solid transparent; border-radius: 12px; font-size: 15px;
      font-weight: 600; text-decoration: none; cursor: pointer; transition: all .25s;
      display: inline-flex; align-items: center; gap: 8px; background: #f2f4f8; color: #4a5568;
      box-shadow: 0 1px 3px rgba(0,0,0,.05);
    }
    .tool-nav-btn:hover {
      border-color: #4472c4; color: #4472c4; background: #eef3ff;
      box-shadow: 0 3px 12px rgba(68,114,196,.18); transform: translateY(-2px);
    }
    .tool-nav-btn.active {
      background: linear-gradient(135deg, #4472c4 0%, #618cef 100%);
      color: #fff; border-color: #4472c4; pointer-events: none;
      box-shadow: 0 4px 16px rgba(68,114,196,.35);
    }
    .usage-card {
      background: #fff; border-radius: 10px; padding: 16px 20px; margin-bottom: 16px;
      box-shadow: 0 1px 4px rgba(0,0,0,.08); border-left: 4px solid #4472c4;
    }
    .usage-card .usage-title { font-size: 14px; font-weight: 600; color: #333; margin-bottom: 8px; }
    .usage-card .usage-text { font-size: 13px; color: #555; line-height: 1.8; }
  </style>
</head>
<body>
  <div class="container">
    <h1>Jira 评论转单分析工具</h1>
    <p class="subtitle">输入账号密码和 Bug ID，自动分析备注中的转单/提单关键词</p>

    <!-- 工具导航 -->
    <div class="tool-nav">
      <a class="tool-nav-btn" href="/">🔍 Bug 状态查询</a>
      <a class="tool-nav-btn" href="/huawei">🔗 华为链接检查</a>
      <span class="tool-nav-btn active">📋 评论转单分析</span>
    </div>

    <!-- 使用场景说明 -->
    <div class="usage-card">
      <div class="usage-title">📝 使用场景说明</div>
      <div class="usage-text">本页面主要在实际工作中，在手机侧名下的问题，需要重新审视，此时需要人肉找到原有的备注信息中哪位互联开发分析提单，此工具根据关键词/关键语境批量分析备注信息的内容，最后找到原始开发，无需人肉处理</div>
    </div>

    <!-- 配置卡片 -->
    <div class="card">
      <h2>配置信息</h2>
      <div class="form-row">
        <div class="form-group" style="max-width:320px">
          <label>Jira 服务地址</label>
          <input type="text" id="jiraUrl" value="https://jira.gwm.cn">
        </div>
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>用户名</label>
          <input type="text" id="username" placeholder="GW00295409" value="GW00295409">
        </div>
        <div class="form-group">
          <label>密码</label>
          <input type="password" id="password" placeholder="Jira 密码" value="5352Zhao123456.">
        </div>
      </div>
      <div class="form-row">
        <div class="form-group">
          <label>Bug ID 列表</label>
          <textarea id="bugList" placeholder="SCHL-5918&#10;SCHL-3731&#10;支持逗号、空格、换行分隔"></textarea>
          <div class="hint">多个 ID 用逗号、空格或换行分隔</div>
        </div>
      </div>
      <div class="btn-row">
        <button class="btn btn-primary" id="btnStart" onclick="startAnalysis()">开始分析</button>
        <button class="btn btn-secondary" onclick="document.getElementById('log').classList.toggle('show')">显示/隐藏日志</button>
      </div>
      <div class="progress-bar" id="progressBar"><div class="fill" id="progressFill"></div></div>
      <div class="log" id="log"></div>
    </div>

    <!-- 结果卡片 -->
    <div id="resultCard" class="card" style="display:none">
      <h2>分析结果</h2>
      <div class="stats" id="stats"></div>
      <div class="btn-row">
        <button class="btn btn-export" id="btnExport" onclick="exportExcel()">导出 Excel (.xlsx)</button>
        <span id="resultTime" style="font-size:12px;color:#aaa"></span>
      </div>
      <div id="authorSection" style="margin-top:12px"></div>
      <div style="overflow-x:auto;max-height:500px;overflow-y:auto;margin-top:10px" id="tableWrap"></div>
    </div>
  </div>

  <div class="toast" id="toast"></div>

  <script>
    let allResults = [];

    function log(msg, type) {
      const el = document.getElementById('log');
      el.classList.add('show');
      el.innerHTML += '<div class="' + type + '">' + msg + '</div>';
      el.scrollTop = el.scrollHeight;
    }
    function showToast(msg, type) {
      const t = document.getElementById('toast');
      t.textContent = msg; t.className = 'toast toast-' + type + ' show';
      setTimeout(() => t.className = 'toast', 3000);
    }
    function setProgress(pct) {
      document.getElementById('progressBar').style.display = pct > 0 ? 'block' : 'none';
      document.getElementById('progressFill').style.width = pct + '%';
    }
    function escHtml(s) {
      if (!s) return '';
      return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
    }

    async function startAnalysis() {
      const url = document.getElementById('jiraUrl').value.trim().replace(/\/+$/, '');
      const username = document.getElementById('username').value.trim();
      const password = document.getElementById('password').value.trim();
      const bugText = document.getElementById('bugList').value.trim();
      if (!url || !username || !password) { showToast('请填写完整的配置信息', 'error'); return; }
      if (!bugText) { showToast('请输入至少一个 Bug ID', 'error'); return; }
      const ids = bugText.split(/[,，\s\n]+/).map(s => s.trim()).filter(Boolean);
      if (!ids.length) { showToast('未检测到有效 Bug ID', 'error'); return; }

      const btn = document.getElementById('btnStart');
      btn.disabled = true; btn.textContent = '分析中...';
      document.getElementById('resultCard').style.display = 'none';
      document.getElementById('log').innerHTML = ''; allResults = []; setProgress(0);
      log('开始分析 ' + ids.length + ' 个 Issue...', 'info');

      for (let i = 0; i < ids.length; i++) {
        const id = ids[i];
        setProgress(Math.round((i / ids.length) * 100));
        log('--- [' + id + '] (' + (i + 1) + '/' + ids.length + ') ---', 'info');
        try {
          const resp = await fetch('/api/transfer/check', {
            method: 'POST', headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ jira_url: url, username, password, issue_id: id })
          });
          const data = await resp.json();
          if (data.error) { log('  ' + data.error, 'err'); }
          else if (data.hits && data.hits.length > 0) {
            log('  命中 ' + data.hits.length + ' 条转单/提单备注', 'ok');
            allResults = allResults.concat(data.hits);
          } else { log('  未发现转单/提单类备注', 'info'); }
        } catch (e) { log('  请求失败: ' + e.message, 'err'); }
      }

      setProgress(100);
      btn.disabled = false; btn.textContent = '开始分析';
      log('分析完成！共命中 ' + allResults.length + ' 条', allResults.length > 0 ? 'ok' : 'info');

      if (allResults.length > 0) { renderResults(); showToast('分析完成，发现 ' + allResults.length + ' 条结果', 'success'); }
      else { showToast('分析完成，未发现转单/提单类备注', 'error'); }
    }

    function renderResults() {
      document.getElementById('resultCard').style.display = 'block';

      const authorCount = {};
      allResults.forEach(r => { authorCount[r.author_display] = (authorCount[r.author_display] || 0) + 1; });

      document.getElementById('stats').innerHTML =
        '<div class="stat-item"><div class="num">' + allResults.length + '</div><div class="label">命中总数</div></div>' +
        '<div class="stat-item"><div class="num">' + Object.keys(authorCount).length + '</div><div class="label">涉及作者</div></div>';

      let authorHtml = '<div style="display:flex;gap:10px;flex-wrap:wrap">';
      Object.entries(authorCount).sort((a,b) => b[1] - a[1]).forEach(([name, count]) => {
        authorHtml += '<div class="author-card"><div class="count">' + count + '</div><div class="name">' + escHtml(name) + '</div></div>';
      });
      authorHtml += '</div>';
      document.getElementById('authorSection').innerHTML = authorHtml;

      let html = '<table><thead><tr>' +
        '<th>Issue ID</th><th>Jira链接</th><th>备注ID</th><th>备注时间</th><th>作者</th><th>匹配关键词</th><th>备注摘要</th></tr></thead><tbody>';
      allResults.forEach(r => {
        html += '<tr>' +
          '<td><b>' + r.issue_id + '</b></td>' +
          '<td><a href="' + r.jira_url + '" target="_blank" style="color:#4472c4;text-decoration:none">' + r.issue_id + '</a></td>' +
          '<td>' + r.comment_id + '</td>' +
          '<td style="white-space:nowrap">' + r.created + '</td>' +
          '<td>' + r.author_display + '</td>' +
          '<td style="color:#c77d00;font-weight:500">' + escHtml(r.matched_keywords) + '</td>' +
          '<td><span style="color:#666;font-size:11px" title="' + escHtml(r.comment_body_plain) + '">' + escHtml((r.comment_body_plain || '').substring(0, 80)) + '...</span></td>' +
        '</tr>';
      });
      html += '</tbody></table>';
      document.getElementById('tableWrap').innerHTML = html;
      document.getElementById('resultTime').textContent = '分析时间: ' + new Date().toLocaleString('zh-CN');
    }

    function exportExcel() {
      if (!allResults.length) return;
      const form = document.createElement('form');
      form.method = 'POST'; form.action = '/api/transfer/export';
      const input = document.createElement('input');
      input.type = 'hidden'; input.name = 'data'; input.value = JSON.stringify(allResults);
      form.appendChild(input);
      document.body.appendChild(form); form.submit(); document.body.removeChild(form);
    }
  </script>
</body>
</html>"""


# ─────────────────────────────────────────────
# HTTP 处理器
# ─────────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):

    def log_message(self, format, *args):
        pass  # 静默服务器日志

    def send_json(self, data, code=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        if self.path in ("/", "/index.html"):
            body = HTML_PAGE.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        elif self.path == "/api/config":
            # 返回服务端配置
            cfg = load_config()
            self.send_json({
                "jira_base_url": cfg.get("jira_base_url", "https://jira.gwm.cn"),
                "jira_username": cfg.get("jira_username", ""),
                "jira_password": cfg.get("jira_password", ""),
                "dingtalk_webhook": cfg.get("dingtalk_webhook", ""),
            })

        elif self.path == "/huawei":
            body = HUAWEI_HTML_PAGE.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        elif self.path == "/transfer":
            body = TRANSFER_HTML_PAGE.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length)
        try:
            payload = json.loads(body.decode("utf-8"))
        except Exception:
            self.send_json({"error": "JSON 解析失败"}, 400)
            return

        # ── /api/status ── 查询单个 Issue
        if self.path == "/api/status":
            issue_id = payload.get("issue_id", "").strip().upper()
            if not issue_id:
                self.send_json({"error": "缺少 issue_id"}, 400)
                return
            username = payload.get("username", "").strip()
            password = payload.get("password", "")
            if not username or not password:
                self.send_json({"error": "缺少用户名或密码"}, 400)
                return
            cfg         = load_config()
            base_url    = cfg.get("jira_base_url", DEFAULT_CONFIG["jira_base_url"]).rstrip("/")
            auth_header = make_auth_header(username, password)
            result      = get_issue_status(issue_id, base_url, auth_header)
            self.send_json(result)

        # ── /api/notify ── 钉钉批量通知
        elif self.path == "/api/notify":
            results      = payload.get("results", [])
            webhook_url  = payload.get("webhook", "").strip()
            custom_text  = payload.get("custom_text", "").strip()

            if not results:
                self.send_json({"success": False, "error": "没有查询结果，请先查询状态"}, 400)
                return
            if not webhook_url:
                cfg = load_config()
                webhook_url = cfg.get("dingtalk_webhook", "").strip()
            if not webhook_url:
                self.send_json({"success": False, "error": "未配置钉钉 Webhook URL"}, 400)
                return

            cfg      = load_config()
            base_url = cfg.get("jira_base_url", DEFAULT_CONFIG["jira_base_url"]).rstrip("/")
            result   = send_dingtalk_notify(webhook_url, results, base_url, custom_text or None)
            self.send_json(result)

        # ── /api/preview-msg ── 生成消息预览
        elif self.path == "/api/preview-msg":
            results = payload.get("results", [])
            if not results:
                self.send_json({"success": False, "error": "无数据"}, 400)
                return
            cfg      = load_config()
            base_url = cfg.get("jira_base_url", DEFAULT_CONFIG["jira_base_url"]).rstrip("/")
            text, open_count, assignee_count, assignees, _ = build_markdown_text(results, base_url)
            self.send_json({
                "success": True,
                "text": text,
                "open_count": open_count,
                "assignee_count": assignee_count,
                "assignees": assignees,
            })

        # ── /api/save-webhook ── 保存 Webhook 到 config.json
        elif self.path == "/api/save-webhook":
            webhook = payload.get("webhook", "").strip()
            if not webhook:
                self.send_json({"success": False, "error": "webhook 为空"}, 400)
                return
            try:
                cfg = load_config()
                cfg["dingtalk_webhook"] = webhook
                with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                    json.dump(cfg, f, ensure_ascii=False, indent=2)
                self.send_json({"success": True})
            except Exception as e:
                self.send_json({"success": False, "error": str(e)})

        # ── /api/export ── 导出 Excel
        elif self.path == "/api/export":
            results = payload.get("results", [])
            if not results:
                self.send_json({"error": "无数据"}, 400)
                return
            try:
                import importlib
                openpyxl = importlib.import_module("openpyxl")
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "【首页】手车互联PM提效工具"

                headers = ["Jira ID", "状态", "优先级", "经办人", "解决方案", "摘要"]
                header_fill = PatternFill("solid", fgColor="1E3A5F")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                thin_border = Border(
                    left=Side(style="thin", color="DDDDDD"),
                    right=Side(style="thin", color="DDDDDD"),
                    top=Side(style="thin", color="DDDDDD"),
                    bottom=Side(style="thin", color="DDDDDD"),
                )
                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=h)
                    cell.fill      = header_fill
                    cell.font      = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = thin_border

                base_url = load_config().get("jira_base_url", "https://jira.gwm.cn").rstrip("/")
                for row_idx, r in enumerate(results, 2):
                    row_data = [
                        r.get("id", ""),
                        r.get("error") or r.get("status", ""),
                        r.get("priority", "—"),
                        r.get("assignee", "—"),
                        r.get("resolution", "—"),
                        r.get("summary", "—"),
                    ]
                    status   = r.get("status", "")
                    if r.get("error"):
                        row_fill = PatternFill("solid", fgColor="FDECEA")
                    elif status in DONE_STATUSES:
                        row_fill = PatternFill("solid", fgColor="E8F8EF")
                    elif row_idx % 2 == 0:
                        row_fill = PatternFill("solid", fgColor="F7F9FC")
                    else:
                        row_fill = None

                    for col_idx, val in enumerate(row_data, 1):
                        cell           = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.border    = thin_border
                        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 6))
                        if row_fill:
                            cell.fill = row_fill
                        if col_idx == 1 and not r.get("error"):
                            cell.hyperlink = f"{base_url}/browse/{r['id']}"
                            cell.font      = Font(color="2D6CDF", underline="single")

                ws.column_dimensions["A"].width = 14
                ws.column_dimensions["B"].width = 14
                ws.column_dimensions["C"].width = 10
                ws.column_dimensions["D"].width = 14
                ws.column_dimensions["E"].width = 12
                ws.column_dimensions["F"].width = 50
                ws.row_dimensions[1].height     = 22
                ws.freeze_panes                 = "A2"
                ws.auto_filter.ref              = f"A1:F{len(results)+1}"

                out_path = os.path.join(APP_DIR, "jira_status_result.xlsx")
                wb.save(out_path)
                self.send_json({"success": True, "path": out_path})
            except ImportError:
                self.send_json({"error": "缺少 openpyxl，请运行：pip install openpyxl"})
            except Exception as e:
                self.send_json({"error": str(e)})

        # ── /api/huawei/check ── 华为链接检查
        elif self.path == "/api/huawei/check":
            base_url = payload.get("jira_url", "").rstrip("/")
            username = payload.get("username", "")
            password = payload.get("password", "")
            issue_id = payload.get("issue_id", "")
            check_linked = payload.get("check_linked", True)

            if not all([base_url, username, password, issue_id]):
                self.send_json({"error": "参数不完整"}, 400)
                return

            auth_header = make_auth(username, password)
            result = check_issue(issue_id, base_url, auth_header, check_linked)
            self.send_json(result)

        # ── /api/huawei/export ── 华为链接结果导出
        elif self.path == "/api/huawei/export":
            try:
                import urllib.parse
                if self.headers.get("Content-Type", "").startswith("application/json"):
                    rows = payload.get("data", [])
                else:
                    pairs = urllib.parse.parse_qs(body.decode("utf-8"))
                    data_str = pairs.get("data", ["[]"])[0]
                    rows = json.loads(data_str)
            except Exception as e:
                self.send_json({"error": f"解析失败: {e}"}, 400)
                return

            if not rows:
                self.send_json({"error": "无数据可导出"}, 400)
                return

            try:
                file_data, file_ext = build_huawei_xlsx(rows)
                mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                filename = "jira_huawei_links_result.xlsx"
            except ImportError:
                file_data, file_ext = build_huawei_csv(rows)
                mime = "text/csv; charset=utf-8"
                filename = "jira_huawei_links_result.csv"

            self.send_response(200)
            self.send_header("Content-Type", mime)
            self.send_header("Content-Disposition", f"attachment; filename=\"{filename}\"")
            self.send_header("Content-Length", str(len(file_data)))
            self.end_headers()
            self.wfile.write(file_data)

        # ── /api/transfer/check ── 评论转单分析
        elif self.path == "/api/transfer/check":
            base_url = payload.get("jira_url", "").rstrip("/")
            username = payload.get("username", "")
            password = payload.get("password", "")
            issue_id = payload.get("issue_id", "")

            if not all([base_url, username, password, issue_id]):
                self.send_json({"error": "参数不完整"}, 400)
                return

            auth_header = make_auth(username, password)
            hits, err = get_transfer_comments(issue_id, base_url, auth_header)
            if err:
                self.send_json({"error": err})
            else:
                self.send_json({"hits": hits})

        # ── /api/transfer/export ── 评论转单结果导出
        elif self.path == "/api/transfer/export":
            try:
                if self.headers.get("Content-Type", "").startswith("application/json"):
                    rows = payload.get("data", [])
                else:
                    pairs = urllib.parse.parse_qs(body.decode("utf-8"))
                    data_str = pairs.get("data", ["[]"])[0]
                    rows = json.loads(data_str)
            except Exception as e:
                self.send_json({"error": f"解析失败: {e}"}, 400)
                return

            if not rows:
                self.send_json({"error": "无数据可导出"}, 400)
                return

            try:
                file_data, file_ext = build_transfer_xlsx(rows)
                mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                filename = "jira_transfer_analysis_result.xlsx"
            except ImportError:
                file_data, file_ext = build_transfer_csv(rows)
                mime = "text/csv; charset=utf-8"
                filename = "jira_transfer_analysis_result.csv"

            self.send_response(200)
            self.send_header("Content-Type", mime)
            self.send_header("Content-Disposition", f"attachment; filename=\"{filename}\"")
            self.send_header("Content-Length", str(len(file_data)))
            self.end_headers()
            self.wfile.write(file_data)

        else:
            self.send_response(404)
            self.end_headers()


# ─────────────────────────────────────────────
# 启动
# ─────────────────────────────────────────────
def main():
    # 允许端口重用，防止上次异常退出后端口被占用
    socketserver.TCPServer.allow_reuse_address = True

    # 尝试启动，端口被占用则自动杀掉旧进程
    used_port = PORT
    for attempt in range(5):
        try:
            server = HTTPServer(("127.0.0.1", used_port), Handler)
            break
        except OSError:
            # 杀掉占用端口的旧进程
            import subprocess, platform
            if platform.system() == "Windows":
                result = subprocess.run(
                    ["netstat", "-ano", "|", "findstr", f":{used_port}"],
                    capture_output=True, text=True, shell=True
                )
                for line in result.stdout.strip().split("\n"):
                    parts = line.split()
                    if parts:
                        pid = parts[-1]
                        subprocess.run(["taskkill", "/F", "/PID", pid], capture_output=True)
            else:
                result = subprocess.run(["lsof", "-ti", f":{used_port}"], capture_output=True, text=True)
                for pid in result.stdout.strip().split("\n"):
                    if pid:
                        subprocess.run(["kill", "-9", pid], capture_output=True)
            if attempt == 0:
                continue  # 杀了旧进程后重试同端口
            used_port = PORT + attempt
    else:
        print(f"❌ 端口 {PORT}-{PORT+4} 全部被占用，无法启动")
        return

    url = f"http://localhost:{used_port}"
    print("=" * 55)
    print(f"  🛠️  手车互联 PM 提效工具集  已启动")
    print(f"  👉 浏览器访问：{url}")
    print(f"  📌  Bug 状态查询  →  {url}")
    print(f"  📌  华为链接检查  →  {url}/huawei")
    print(f"  📌  评论转单分析  →  {url}/transfer")
    print(f"  按 Ctrl+C 停止服务")
    print("=" * 55)
    threading.Timer(1.2, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n✋ 服务已停止")
        server.server_close()

if __name__ == "__main__":
    # 确保退出时清理端口
    import atexit, signal, platform
    def cleanup(*_):
        import subprocess
        if platform.system() == "Windows":
            result = subprocess.run(
                ["netstat", "-ano", "|", "findstr", f":{PORT}"],
                capture_output=True, text=True, shell=True
            )
            for line in result.stdout.strip().split("\n"):
                parts = line.split()
                if parts:
                    pid = parts[-1]
                    subprocess.run(["taskkill", "/F", "/PID", pid], capture_output=True)
        else:
            result = subprocess.run(["lsof", "-ti", f":{PORT}"], capture_output=True, text=True)
            for pid in result.stdout.strip().split("\n"):
                if pid:
                    subprocess.run(["kill", "-9", pid], capture_output=True)
    atexit.register(cleanup)
    signal.signal(signal.SIGTERM, cleanup)
    main()
